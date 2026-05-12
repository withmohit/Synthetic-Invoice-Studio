import random
from datetime import datetime

import openpyxl
import numpy as np
from faker import Faker

faker = Faker()


def infer_cell_value(sample_value=None):
    def numeric(min_value=1, max_value=1000, decimals=0):
        if decimals == 0:
            return random.randint(min_value, max_value)
        return round(random.uniform(min_value, max_value), decimals)

    def is_int_string(value):
        try:
            int(value)
            return True
        except ValueError:
            return False

    def is_float_string(value):
        try:
            float(value)
            return True
        except ValueError:
            return False

    if isinstance(sample_value, bool):
        return lambda: random.choice([True, False])

    if isinstance(sample_value, int):
        return lambda: numeric(1, max(1, abs(sample_value) * 5), decimals=0)

    if isinstance(sample_value, float):
        return lambda: numeric(1, max(1, abs(int(sample_value) * 5)), decimals=2)

    if isinstance(sample_value, datetime):
        return lambda: faker.date_between(start_date='-2y', end_date='today').strftime('%m.%Y')

    if isinstance(sample_value, str):
        value = sample_value.strip()
        if value == '':
            return lambda: ''
        if is_int_string(value):
            int_value = int(value)
            return lambda: numeric(1, max(1, abs(int_value) * 5), decimals=0)
        if is_float_string(value):
            float_value = float(value)
            return lambda: numeric(1, max(1, abs(int(float_value)) * 5), decimals=2)
        parts = value.split()
        if len(parts) == 1:
            return lambda: faker.word().title()
        if len(parts) <= 3:
            return lambda: ' '.join(faker.words(nb=random.randint(1, 3))).title()
        return lambda: faker.sentence(nb_words=random.randint(3, 7))

    return lambda: faker.sentence(nb_words=random.randint(2, 5))


def generate_synthetic_excel(path='synthetic_invoice.xlsx', template_path='test_invoice.xlsx', total_rows=80):
    source_data = read_excel_to_data(template_path)
    bbox = detect_table_bbox(source_data)
    if bbox is None:
        raise ValueError('Unable to detect a dense table in the template workbook.')

    start_row, end_row, start_col, end_col = bbox
    headers = [source_data[start_row][col] for col in range(start_col, end_col + 1)]

    sample_row = None
    for row in source_data[start_row + 1:end_row + 1]:
        row_slice = row[start_col:end_col + 1]
        if any(cell is not None and str(cell).strip() for cell in row_slice):
            sample_row = row_slice
            break
    if sample_row is None:
        sample_row = headers

    generators = [infer_cell_value(sample) for sample in sample_row]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'SyntheticInvoice'

    sheet.append(headers)
    for _ in range(total_rows - 1):
        sheet.append([gen() for gen in generators])

    workbook.save(path)
    return path


def read_excel_to_data(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    return [row for row in sheet.iter_rows(values_only=True)]


def detect_table_bbox(data, min_row_density=0.3):
    temp = [[str(cell).strip() if cell is not None else '' for cell in row] for row in data]
    mask = np.array([[bool(cell) for cell in row] for row in temp])

    if mask.size == 0:
        return None

    row_density = mask.sum(axis=1) / mask.shape[1]
    dense_rows = row_density >= min_row_density

    if not dense_rows.any():
        return None

    blocks = []
    start = None
    for i, is_dense in enumerate(dense_rows):
        if is_dense and start is None:
            start = i
        elif not is_dense and start is not None:
            blocks.append((start, i - 1))
            start = None

    if start is not None:
        blocks.append((start, len(dense_rows) - 1))

    start_row, end_row = max(blocks, key=lambda x: x[1] - x[0])
    submask = mask[start_row:end_row + 1, :]
    col_density = submask.sum(axis=0) / submask.shape[0]
    dense_cols = col_density >= min_row_density

    if not dense_cols.any():
        return None

    col_blocks = []
    start = None
    for i, is_dense in enumerate(dense_cols):
        if is_dense and start is None:
            start = i
        elif not is_dense and start is not None:
            col_blocks.append((start, i - 1))
            start = None

    if start is not None:
        col_blocks.append((start, len(dense_cols) - 1))

    start_col, end_col = max(col_blocks, key=lambda x: x[1] - x[0])
    print(f'Table bounding box: Rows {start_row}-{end_row}, Columns {start_col}-{end_col}')
    return start_row, end_row, start_col, end_col


if __name__ == '__main__':
    output_path = generate_synthetic_excel('synthetic_invoice_dynamic.xlsx', total_rows=80)
    print(f'Generated Excel file: {output_path} at {datetime.now():%Y-%m-%d %H:%M:%S}')

    data = read_excel_to_data(output_path)
    print('First 10 rows:')
    for row in data[:10]:
        print(row)

    bbox = detect_table_bbox(data)
    print('Detected bbox:', bbox)